import pandas as pd
import csv, os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from AutoRecordTimeSheet import ControlManage

def AutoGetMonthData(csvPath):
    MonthDataDict = {}
    with open(csvPath, mode='r', newline='', encoding='utf-8') as csvfile:
        csvReader = csv.reader(csvfile)
        for row in csvReader:
            if "Date" in row:
                continue
            key = int(row[0].split("/")[2])
            infoList = []
            if ":" in row[1]:
                sHour = int(row[1].split(":")[0])
                sMinute = int(row[1].split(":")[1])
                infoList.append(sHour)
                infoList.append(sMinute)
            else:
                infoList.append(-1)
                infoList.append(-1)
            if ":" in row[2]:
                eHour = int(row[2].split(":")[0])
                eMinute = int(row[2].split(":")[1])
                infoList.append(eHour)
                infoList.append(eMinute)
            else:
                infoList.append(-1)
                infoList.append(-1)

            MonthDataDict[key] = infoList
    return MonthDataDict


def AutoAttendenceQuery(MonthDataDict, xwlPath, startRow, startHCol, endHCol):
    workbook = load_workbook(xwlPath)
    sheet = workbook.active
    for key, val in MonthDataDict.items():
        rowLineVal = startRow + int(key) -1
        sH = sheet.cell(row=rowLineVal, column=startHCol)
        sM = sheet.cell(row=rowLineVal, column=startHCol + 1)
        eH = sheet.cell(row=rowLineVal, column=endHCol)
        eM = sheet.cell(row=rowLineVal, column=endHCol + 1)
        sH.value = val[0]
        sM.value = val[1]
        eH.value = val[2]
        eM.value = val[3]
        
        startContMerge = 'Y' + str(rowLineVal)
        mergeContCell = startContMerge + ':' + 'AE' + str(rowLineVal)
        sheet.merge_cells(mergeContCell)
        sheet[startContMerge] = "ETCとVICS開発"
        #sheet[startMerge].alignment = Alignment(horizontal='center', vertical='center')            

        startAttMerge = 'E' + str(rowLineVal)
        mergeAttCell = startAttMerge + ':' + 'H' + str(rowLineVal)
        sheet.merge_cells(mergeAttCell)
        sheet[startAttMerge] = "出勤（Bosch）"
        sheet[startAttMerge].alignment = Alignment(horizontal='center', vertical='center')   

    return workbook

def RunningQuerySheet():
    fileName = ControlManage()
    filePath = "C:\\worksrc\\PythonCode\\" + fileName
    monthDict = AutoGetMonthData(filePath)
    orgPath = "C:\\Users\\a1294.zhou\\WorkingFiles\\Doc_Images\\考勤相关\\Quefy株式会社2024年10月作業報告書(周尚軍).xlsx"
    startRow = 18
    startHCol = 9
    endHCol = 11
    typeQuery = 1
    workbook1 = AutoAttendenceQuery(monthDict, orgPath, startRow, startHCol, endHCol)
    curScriptPath = os.path.dirname(os.path.abspath(__file__))
    newXlsxName = curScriptPath + orgPath.split("\\")[-1]
    workbook1.save(newXlsxName)

#RunningQuerySheet()

def AutoAttendenceUniq(MonthDataDict, xwlPath, startRow):
    workbook = load_workbook(xwlPath)
    sheet = workbook.active
    for key, val in MonthDataDict.items():
        rowLineVal = startRow + int(key) -1
        
        sHour = str(val[0]).zfill(2) + ":" + str(val[1]).zfill(2)
        eHour = str(val[2]).zfill(2) + ":" + str(val[3]).zfill(2)

        startHourMerge = 'F' + str(rowLineVal)
        mergeSatrtCell = startHourMerge + ':' + 'G' + str(rowLineVal)
        sheet.merge_cells(mergeSatrtCell)
        sheet[startHourMerge] = sHour
        sheet[startHourMerge].alignment = Alignment(horizontal='center', vertical='center')   

        endHourMerge = 'H' + str(rowLineVal)
        mergeEndCell = endHourMerge + ':' + 'I' + str(rowLineVal)
        sheet.merge_cells(mergeEndCell)
        sheet[endHourMerge] = eHour
        sheet[endHourMerge].alignment = Alignment(horizontal='center', vertical='center')   

        startRestMerge = 'J' + str(rowLineVal)
        mergeRestCell = startRestMerge + ':' + 'K' + str(rowLineVal)
        sheet.merge_cells(mergeRestCell)
        sheet[startRestMerge] = "01:00"
        sheet[startRestMerge].alignment = Alignment(horizontal='center', vertical='center')   

        # deMerge = 'D' + str(rowLineVal)
        # mergeDECell = deMerge + ':' + 'E' + str(rowLineVal)
        # sheet.merge_cells(mergeDECell)
        # sheet[deMerge] = "出勤"
        # sheet[deMerge].alignment = Alignment(horizontal='center', vertical='center')

        tuMerge = 'T' + str(rowLineVal)
        mergeTUCell = tuMerge + ':' + 'U' + str(rowLineVal)
        sheet.merge_cells(mergeTUCell)
        sheet[tuMerge] = "出勤"
        sheet[tuMerge].alignment = Alignment(horizontal='center', vertical='center')

    return workbook

def RunningUniqSheet():
    fileName = ControlManage()
    filePath = "C:\\worksrc\\PythonCode\\" + fileName
    monthDict = AutoGetMonthData(filePath)
    orgPath = "C:\\Users\\a1294.zhou\\WorkingFiles\\Doc_Images\\考勤相关\\202410_作業時間内訳表_周尚軍.xlsx"
    startRow = 10
    workbook1 = AutoAttendenceUniq(monthDict, orgPath, startRow)
    curScriptPath = os.path.dirname(os.path.abspath(__file__))
    newXlsxName = curScriptPath + orgPath.split("\\")[-1]
    workbook1.save(newXlsxName)

RunningUniqSheet()

#RunningQuerySheet()